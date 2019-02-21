# 该方法为调用openpyxl的方法，速度较慢，耗时约3分钟，但是会保留样式
from openpyxl import load_workbook
import os


# 查找关键字方法
def find_keyword_in_sheet(sheet, keyword):
    for i, cell2 in enumerate(sheet['C']):
        try:
            if cell2:
                if str(keyword) in str(cell2.value):
                    return i
        except Exception as e:
            break
    return None


# 计算超额费用 参数为：总费用、限额
def get_bonus(total, limit):
    # 总费用减去限额小于0得0，大于0取差
    return 0 if total - limit <= 0 else round(total - limit, 2)


# 获取额度和超额费用 参数为：套餐费、额度、基础费、总费用
def get_limit_and_bonus(raw, raw_limit, base, total):
    limit = 0
    bonus = 0
    # 套餐费为36，总费用小于60，限额为36，超额为0
    if raw == 36 and total < 60:
        limit = 36
        bonus = 0
    # 套餐费为36，基础费大于60，限额为63，超额为差额计算（先取额度大的数）
    elif raw == 36 and base >= 60:
        limit = 63
        if raw_limit > limit:
            limit = raw_limit
        bonus = get_bonus(total, limit)
    # 套餐费为60，基础费等于60，限额为60，超额为差额计算（先取额度大的数）
    elif raw == 60 and base == 60:
        limit = 60
        if raw_limit > limit:
            limit = raw_limit
        bonus = get_bonus(total, limit)
    # 套餐费为60，基础费在60与70之间，限额为87，超额为差额计算（先取额度大的数）
    elif raw == 60 and 60 < base <= 70:
        limit = 87
        if raw_limit > limit:
            limit = raw_limit
        bonus = get_bonus(total, limit)
    # 套餐费为60，基础费在70之上，限额为90，超额为差额计算（先取额度大的数）
    elif raw == 60 and base > 70:
        limit = 90
        if raw_limit > limit:
            limit = raw_limit
        bonus = get_bonus(total, limit)
    # 套餐费大于60，超额为差额计算（先取额度大的数）
    elif raw > 60:
        limit = base
        if raw_limit > limit:
            limit = raw_limit
        bonus = get_bonus(total, limit)
    return limit, bonus


# 拆分合并单元格
def unmerge_cells(worksheet):
    # 注意这里为什么要用循环，因为拆分单元格会使之前合并的单元格行号减1，从而使被改变行号的合并单元格无法拆分
    while True:
        # 合并单元格的位置信息，可迭代对象（单个是一个'openpyxl.worksheet.cell_range.CellRange'对象），print后就是excel坐标信息
        m_list = worksheet.merged_cells
        break_flag = 0
        for m_area in m_list:
            # 合并单元格的起始行坐标、终止行坐标。。。。，
            r1, r2, c1, c2 = m_area.min_row, m_area.max_row, m_area.min_col, m_area.max_col
            # 这里注意需要把合并单元格的信息提取出再拆分
            worksheet.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
            # print('符合条件%s' % str(m_area))
            break_flag += 1
        if break_flag == 0:
            break


# 读取Excel
def read_excel():
    print('请确保同文件夹中存在包含‘扣款’，‘话单’的两个文件')
    file_path_A = ''
    file_path_B = ''
    # 获取文件路径
    # 利用 os 模块中的 listdir 函数，将路径中的所有文件存储到一个 list 变量中。
    files = os.listdir('.')
    # 利用 for 语句浏览 list 变量中的所有元素
    flag_a = False  # 设置标志位，只查找第一个
    flag_b = False
    for f in files:
        # 利用 if 语句判断文件名是否符合要求。其中， endswitch 函数用来判断一个字符串是否包含某个后缀。
        # 成员运算符 in 用来判断一个字符串是否包含某个子串。不同的条件用 and 或者 or 来连接。
        if file_path_A != '' and file_path_B != '':
            break
        if '扣款' in f and not flag_a:
            print('找到 ' + f)
            file_path_A = f
            flag_a = True
        if '话单' in f and not flag_b:
            print('找到 ' + f)
            file_path_B = f
            flag_b = True
    file_path_C = file_path_A.split('.')[0] + '_程序生成.xlsx'
    file_path_D = '话单未找到号码.txt'
    print('读取文件(%s, %s)...(时间较长，请勿关闭程序)' % (file_path_A, file_path_B))
    # 打开文件
    workbookA = load_workbook(file_path_A)
    workbookB = load_workbook(file_path_B)
    # 猜测格式类型
    workbookA.guess_types = True
    workbookB.guess_types = True
    # 读取表格
    sheetA = workbookA['Sheet1']
    sheetB = workbookB['Sheet1']
    print('读取完成!')
    print('正在计算...')
    # 获取年份月份
    year = ''
    month = ''
    # 获取时间单元格
    try:
        timeCell = sheetB['A2']
        if timeCell:
            year = timeCell.value[:4]       # 年份
            month = timeCell.value[-2:]     # 月份
        else:
            timeCell = sheetB['A3']
            year = timeCell.value[:4]
            month = timeCell.value[-2:]
        print('话单文件A2数据获取：年份为 %s, 月份为 %s' % (year, month))
    except:
        raise RuntimeError('时间格式异常，话单文件A2单元格获取失败')
    # 循环A表
    A_row = 0
    for cellC, cellD, cellE in zip(sheetA['C'], sheetA['D'], sheetA['E']):
        if cellC and A_row != 0:
            # 判断是否为空
            if cellC.value is None:
                break
            # 查找电话号码
            if cellC.value:
                find_B_row = find_keyword_in_sheet(sheetB, cellC.value)
                if find_B_row:
                    raw = cellD.value  # 套餐费
                    raw_limit = int(cellE.value) if str(cellE.value).isdigit() else 0
                    number = str(sheetB['C'][find_B_row].value)  # 电话号码
                    base = sheetB['D'][find_B_row].value  # 基本
                    total = round(sheetB['O'][find_B_row].value, 2)  # 总费用
                    limit, bonus = get_limit_and_bonus(raw, raw_limit, base, total)
                    # sheetA['E'][A_row].value = raw                  # 额度
                    if raw >= 100 and raw_limit != 0:
                        if raw < total < raw_limit:
                            sheetA['F'][A_row].value = total
                        else:
                            sheetA['F'][A_row].value = raw_limit + bonus  # 月总费用
                    else:
                        if raw < total < raw_limit:
                            sheetA['F'][A_row].value = total
                        else:
                            sheetA['F'][A_row].value = raw + bonus  # 月总费用
                    sheetA['G'][A_row].value = bonus  # 超额费用
                    # print(A_row, end='-->')
                    # print(limit, bonus)
                else:
                    print('%s 未找到!' % str(cellC.value))
                    sheetA['F'][A_row].value = '未找到'  # 月总费用
                    sheetA['G'][A_row].value = '未找到'  # 超额费用
        A_row += 1  # 行号加1
    print('计算完成！')

    # 更新固话和宽带数据
    try:
        print('正在更新‘固话’，‘宽带’的列表，(请确保‘扣款’文件中有‘固话’，‘宽带’两张表)，请稍候...')
        # 获取扣款文件中的宽带表
        sheet_broadband = workbookA['宽带']
        # 拆分合并单元格
        unmerge_cells(sheet_broadband)
        # 循环遍历获取
        for i, cellB in enumerate(sheet_broadband['B']):
            # 跳过第一行表头
            if cellB and i != 0:
                # 非空判断
                if cellB.value is None:
                    continue
                broadband_telephone = str(cellB.value)  # 宽带号码
                # 查找话单表获取费用
                j = find_keyword_in_sheet(sheetB, broadband_telephone)
                try:
                    if j:
                        # print('number %s' % broadband_telephone)
                        cost = round(sheetB['O'][j].value, 2)  # 总费用
                        sheet_broadband['D'][i].value = cost
                    else:
                        sheet_broadband['D'][i].value = '未找到'
                except Exception as e:
                    print(broadband_telephone, end='')
                    print(e)
                    continue

        # 获取扣款文件中的固话表
        sheet_fixedline = workbookA['固话']
        # 拆分合并单元格
        unmerge_cells(sheet_fixedline)
        # 确定文件的月份是哪个列
        fixedline_index = 0
        year_month = year + '.' + str(int(month))
        for fixedline_title_cell in list(sheet_fixedline.rows)[0]:
            try:
                if fixedline_title_cell:
                    if str(fixedline_title_cell.value) == year_month:
                        print('找到匹配列 %d' % fixedline_index)
                        break
            except Exception as e:
                print(e)
                continue
            fixedline_index += 1
        # 循环遍历获取
        sheet_fixedline_row = 0
        for cellB in sheet_fixedline['B']:
            # 跳过第一行表头
            if cellB and sheet_fixedline_row != 0:
                try:
                    # 非空判断
                    if cellB.value is None:
                        continue
                    fixedline_telephone = str(cellB.value)  # 固话号码
                    # 查找话单表获取费用
                    j = find_keyword_in_sheet(sheetB, fixedline_telephone)

                    if j:
                        cost = round(sheetB['O'][j].value, 2)  # 总费用
                        sheet_fixedline[sheet_fixedline_row+1][fixedline_index].value = cost
                    else:
                        sheet_fixedline[sheet_fixedline_row+1][fixedline_index].value = '未找到'
                except Exception as e:
                    # print(e)
                    continue
            sheet_fixedline_row += 1
    except Exception as ex:
        print(ex)
        raise ChildProcessError('更新固话、宽带数据失败，请检查是否包含‘固话’‘宽带’表，是否存在 %s 列可用于保存数据' % year_month)

    # 判断文件是否存在，若存在则删除
    if os.path.exists(file_path_C):
        os.remove(file_path_C)
    print('正在保存‘扣款’文件......(时间较长，请勿关闭程序)')
    workbookA.save(file_path_C)
    print('保存成功！请查看文件 %s' % file_path_C)
    # 查询话单文件中的不存在号码
    print('正在匹配话单文件中的号码...')
    f = open(file_path_D, 'w')  # 若是'wb'就表示写二进制文件
    for i, number_cell in enumerate(sheetB['C']):
        try:
            # 非空判断
            if number_cell.value is None or i == 0:
                continue
            number = str(number_cell.value)  # 固话号码
            # 查找话单表获取费用
            j = find_keyword_in_sheet(sheetA, number)
            if not j:
                print('%s 未找到！' % number)
                # sheetB['Q'][i].value = '未找到'
                f.write('%s 未找到\r\n' % number)
        except Exception as e:
            continue
    print('正在保存‘话单未找到号码’文件......(时间较长，请勿关闭程序)')
    f.close()
    print('保存成功！请查看%s' % file_path_D)
    # workbookB.save(file_path_D)


# 主函数
if __name__ == '__main__':
    # read_excel()
    try:
        read_excel()
    except PermissionError as e:
        print('Excel文件被占用了，要先关掉才行')
    except RuntimeError as e:
        print(e)
    except ChildProcessError as e:
        print(e)
    except Exception as e:
        print('请检查（1）是否已将该程序与‘扣款’和‘话单’文件放在一起；（2）并且‘扣款’文件中包含‘固话’，‘宽带’两个表。')
    os.system('pause')
